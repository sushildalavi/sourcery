"""Build the 3 labeling codebooks (Coder A / B / C).

Pipeline:
  1. Load queries_120.json.
  2. For each query, call assistant_answer(scope='uploaded', doc_id=target_doc_id, k=8).
  3. From each answer, extract inline [S#] citations and pair each claim sentence
     with the corresponding citation's evidence chunk (c.snippet).
  4. Repeat for scope='public' (no doc_id pinning) to capture public-mode pairs.
  5. Emit ScholarRAG_Unified_Labeling_Coder_{A,B,C}.xlsx — identical content in
     all three (each coder independently labels the same pairs so we get IAA).

Output xlsx has sheets: Instructions, Rubric, Paper List, Labeling, Summary.

Usage:
    python -m backend.scripts.build_codebooks
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import backend.app as app_module

QUERIES_PATH = Path("Evaluation/queries/queries_120.json")
OUT_DIR = Path("/Users/sushildalavi/Desktop/HUMAN EVAL")

CODERS = ["A", "B", "C"]

# Keep each (claim, evidence) pair row ~this many chars wide in the final
# spreadsheet to avoid overwhelming the reader.
MAX_CLAIM_CHARS = 600
MAX_EVIDENCE_CHARS = 1200

PAPER_LIST = [
    ("01_ResNet.pdf", "Deep Residual Learning for Image Recognition", "Computer vision", "CVPR 2016", "https://arxiv.org/abs/1512.03385"),
    ("02_GAN.pdf", "Generative Adversarial Networks", "Generative models", "NeurIPS 2014", "https://arxiv.org/abs/1406.2661"),
    ("03_Word2Vec.pdf", "Efficient Estimation of Word Representations in Vector Space (Word2Vec)", "Representation learning", "ICLR 2013", "https://arxiv.org/abs/1301.3781"),
    ("04_LLaMA2.pdf", "LLaMA 2: Open Foundation and Fine-Tuned Chat Models", "Open LLMs", "Meta, 2023", "https://arxiv.org/abs/2307.09288"),
    ("05_Chinchilla.pdf", "Training Compute-Optimal Large Language Models (Chinchilla)", "Scaling laws", "NeurIPS 2022", "https://arxiv.org/abs/2203.15556"),
    ("06_ConstitutionalAI.pdf", "Constitutional AI: Harmlessness from AI Feedback", "Alignment", "Anthropic, 2022", "https://arxiv.org/abs/2212.08073"),
    ("07_AlphaGo.pdf", "Mastering the Game of Go with Deep Neural Networks and Tree Search (AlphaGo)", "Reinforcement learning", "Nature 529, 2016", "https://doi.org/10.1038/nature16961"),
    ("08_CLIP.pdf", "Learning Transferable Visual Models From Natural Language Supervision (CLIP)", "Multimodal", "ICML 2021", "https://arxiv.org/abs/2103.00020"),
    ("09_AlphaFold.pdf", "Highly Accurate Protein Structure Prediction with AlphaFold", "Computational biology", "Nature 596, 2021", "https://doi.org/10.1038/s41586-021-03819-2"),
    ("10_StableDiffusion.pdf", "High-Resolution Image Synthesis with Latent Diffusion Models (Stable Diffusion)", "Generative models", "CVPR 2022", "https://arxiv.org/abs/2112.10752"),
    ("11_LSTM.pdf", "Long Short-Term Memory (LSTM)", "Foundational ML", "Neural Computation 9(8), 1997", "https://doi.org/10.1162/neco.1997.9.8.1735"),
    ("12_DQN.pdf", "Human-level Control Through Deep Reinforcement Learning (DQN)", "Reinforcement learning", "Nature 518, 2015", "https://doi.org/10.1038/nature14236"),
    ("13_VAE.pdf", "Auto-Encoding Variational Bayes (VAE)", "Generative models", "ICLR 2014", "https://arxiv.org/abs/1312.6114"),
    ("14_SwinTransformer.pdf", "Swin Transformer: Hierarchical Vision Transformer using Shifted Windows", "Computer vision", "ICCV 2021", "https://arxiv.org/abs/2103.14030"),
    ("15_PageRank.pdf", "The Anatomy of a Large-Scale Hypertextual Web Search Engine (PageRank)", "Information retrieval", "WWW 1998", "http://infolab.stanford.edu/pub/papers/google.pdf"),
]


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def split_sentences(text: str) -> List[str]:
    if not text:
        return []
    parts = re.split(r"(?<=[\.!?])\s+", (text or "").strip())
    return [p.strip() for p in parts if p.strip()]


def extract_pairs(query: Dict, mode: str, resp: Dict) -> List[Dict]:
    """Return one row per cited sentence, pairing it with the cited snippet."""
    answer = (resp or {}).get("answer") or ""
    citations = (resp or {}).get("citations") or []
    if not answer or not citations:
        return []
    cite_by_id = {int(c.get("id") or 0): c for c in citations}
    pairs = []
    for sidx, sentence in enumerate(split_sentences(answer), start=1):
        ids = re.findall(r"\[S?(\d+)\]", sentence)
        if not ids:
            continue
        seen = set()
        for sid_str in ids:
            try:
                sid = int(sid_str)
            except Exception:
                continue
            if sid in seen:
                continue
            seen.add(sid)
            c = cite_by_id.get(sid)
            if not c:
                continue
            snippet = (c.get("snippet") or "").strip()
            if not snippet:
                continue
            claim = re.sub(r"\[(?:S)?\d+\]", "", sentence).strip()
            if not claim:
                continue
            pairs.append(
                {
                    "query_id": query["query_id"],
                    "query": query["query"],
                    "mode": mode,
                    "query_type": query["query_type"],
                    "target_doc_title": query["target_doc_title"],
                    "claim_id": f"{query['query_id']}_s{sidx}_c{sid}",
                    "claim_text": claim[:MAX_CLAIM_CHARS],
                    "evidence_text": snippet[:MAX_EVIDENCE_CHARS],
                }
            )
    return pairs


def run_pipeline(pairs_path: Path, *, include_public: bool = False, per_call_timeout_s: int = 45) -> List[Dict]:
    """Run assistant_answer for each query and dump pairs incrementally.

    Incremental dumping lets us resume/inspect mid-run; per-call timeout
    prevents a single flaky external API from stalling the whole build.
    """
    from concurrent.futures import ThreadPoolExecutor
    from concurrent.futures import TimeoutError as FuturesTimeoutError

    queries = json.loads(QUERIES_PATH.read_text())
    max_q = int(os.getenv("CODEBOOK_MAX_QUERIES", "0") or "0")
    if max_q > 0 and max_q < len(queries):
        # Stratified sample: keep a balanced subset across papers + query types.
        # Round 2 queries are already grouped as 4 types × 2 each × 15 papers;
        # trimming the list uniformly preserves that balance.
        keep_per_type = max(1, max_q // (4 * 15)) if max_q >= 60 else 1
        trimmed: list = []
        for q in queries:
            same_paper_same_type = [
                x for x in trimmed
                if x.get("target_doc_title") == q.get("target_doc_title")
                and x.get("query_type") == q.get("query_type")
            ]
            if len(same_paper_same_type) < keep_per_type:
                trimmed.append(q)
            if len(trimmed) >= max_q:
                break
        queries = trimmed
    print(f"Loaded {len(queries)} queries (include_public={include_public}, cap={max_q or 'none'})")
    all_pairs: List[Dict] = []
    pairs_path.parent.mkdir(parents=True, exist_ok=True)
    pairs_path.write_text("[]")  # start fresh

    modes: List[tuple] = [("uploaded", True)]
    if include_public:
        modes.append(("public", False))

    def _call(payload: Dict) -> Dict:
        return app_module.assistant_answer(payload)

    with ThreadPoolExecutor(max_workers=1) as pool:
        for i, q in enumerate(queries, start=1):
            for mode, is_uploaded in modes:
                payload: Dict = {
                    "query": q["query"],
                    "scope": mode,
                    "k": 8,
                    # Disable general-background public fallback so uploaded-mode
                    # runs stay fully offline (no arXiv / CrossRef hits).
                    "allow_general_background": False,
                }
                if is_uploaded:
                    payload["doc_id"] = q["target_doc_id"]
                fut = pool.submit(_call, payload)
                try:
                    resp = fut.result(timeout=per_call_timeout_s)
                except FuturesTimeoutError:
                    print(f"  [{i:>3}/{len(queries)}] {mode:<8} {q['query_id']:>4}  TIMEOUT after {per_call_timeout_s}s")
                    continue
                except Exception as exc:
                    print(f"  [{i:>3}/{len(queries)}] {mode:<8} {q['query_id']:>4}  FAIL: {type(exc).__name__}: {exc}")
                    continue
                pairs = extract_pairs(q, mode, resp)
                all_pairs.extend(pairs)
                pairs_path.write_text(json.dumps(all_pairs, indent=2, ensure_ascii=False))
                print(
                    f"  [{i:>3}/{len(queries)}] {mode:<8} {q['query_id']:>4}  "
                    f"pairs+={len(pairs):<3}  total={len(all_pairs)}"
                )
    return all_pairs


# ---------------------------------------------------------------------------
# Codebook (xlsx) generation
# ---------------------------------------------------------------------------

_ILLEGAL_XLSX_CHARS = re.compile(
    r"[\000-\010\013\014\016-\037\177]"  # ASCII controls except \t \n \r
)


def _xlsx_safe(value):
    """Strip characters openpyxl refuses to write (PDF extraction leaves these)."""
    if value is None or not isinstance(value, str):
        return value
    return _ILLEGAL_XLSX_CHARS.sub("", value)


def write_codebook(path: Path, coder: str, pairs: List[Dict]) -> None:
    wb = Workbook()

    # 1. Instructions
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = f"ScholarRAG Unified Labeling — Coder {coder}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Coder ID:"
    ws["B3"] = coder
    ws["A4"] = "Total claims:"
    ws["B4"] = len(pairs)
    ws["A5"] = "Your task:"
    ws["B5"] = "Label each claim as 'supported' or 'unsupported' based on the cited evidence."
    ws["A6"] = "Time estimate:"
    ws["B6"] = f"~30-60s per claim · ~{max(1, len(pairs) // 60)} hours total."
    ws["A7"] = "Rule 0:"
    ws["B7"] = "Do NOT share or compare labels with the other coders until everyone is done."

    instructions = [
        "Step-by-step",
        "1. Read the 'Rubric' tab once, top to bottom, before labeling anything.",
        "2. Read the 'Paper List' tab so you know which 15 research papers claims might be about.",
        "3. Open the 'Labeling' tab. Each row = one claim + one piece of evidence.",
        "4. Read col G (claim_text) — identify the specific factual assertion.",
        "5. Read col H (evidence_text) — ask: does this passage actually support that assertion?",
        "6. Click col I (Your Label) and pick 'supported' or 'unsupported' from the dropdown.",
        "7. If uncertain, type a short reason in col J (Notes).",
        "8. Never type free text in col I — use the dropdown only.",
        "9. Do not skip rows. If you genuinely cannot decide, pick the most likely label and explain in Notes.",
        "10. Save the file frequently (Cmd-S / Ctrl-S).",
        "11. The 'Summary' tab shows your progress and label distribution.",
        "12. When done with all rows, save and send the file back.",
        "",
        "Do / Don't",
        "DO  — Use ONLY the evidence passage in col H. No outside knowledge, no Google.",
        "DO  — Label system abstentions ('I only found X in profile/course context...') as UNSUPPORTED.",
        "DO  — For partial matches, if evidence supports MOST of the claim but not all, lean supported and note it.",
        "DO  — Trust your first read. Don't second-guess unless the claim is genuinely ambiguous.",
        "DON'T — Look at other coders' files or discuss labels before everyone finishes.",
        "DON'T — Type anything other than the dropdown options into col I.",
        "DON'T — Skip rows. Every row must have a label before you send the file back.",
    ]
    r = 9
    for line in instructions:
        ws.cell(row=r, column=1, value=line)
        if line in ("Step-by-step", "Do / Don't"):
            ws.cell(row=r, column=1).font = Font(bold=True)
        r += 1
    ws.column_dimensions["A"].width = 110
    ws.column_dimensions["B"].width = 40

    # 2. Rubric
    ws2 = wb.create_sheet("Rubric")
    ws2["A1"] = "Rubric — Supported vs Unsupported"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A3"], ws2["B3"], ws2["C3"] = "Label", "When to use", "Quick test"
    for c in ("A3", "B3", "C3"):
        ws2[c].font = Font(bold=True)
    ws2["A4"] = "supported"
    ws2["B4"] = (
        "Evidence explicitly states or strongly implies the specific factual content of the claim. "
        "A reviewer would accept this passage as the source."
    )
    ws2["C4"] = "Can you point to a sentence in the evidence that a reviewer would accept as grounding for this claim?"
    ws2["A5"] = "unsupported"
    ws2["B5"] = (
        "Evidence does NOT substantiate the claim. Includes: (a) evidence about a different topic/paper; "
        "(b) system abstentions; (c) topically related but does not assert the specific claim; "
        "(d) evidence contradicts the claim."
    )
    ws2["C5"] = "If the claim used this evidence as its only source, would a reviewer call it hallucinated?"

    ws2["A7"] = "Worked examples"
    ws2["A7"].font = Font(bold=True)
    headers = ["Claim", "Evidence snippet", "Correct label", "Why"]
    for i, h in enumerate(headers, start=1):
        ws2.cell(row=8, column=i, value=h).font = Font(bold=True)
    examples = [
        ("Swin Transformer achieves 87.3% top-1 accuracy on ImageNet-1K.",
         "Swin Transformer achieves 87.3% top-1 accuracy on ImageNet-1K validation.",
         "supported",
         "Evidence explicitly states the value."),
        ("GANs are trained with a supervised cross-entropy loss on labeled data.",
         "The framework sets up a minimax game between a generator and a discriminator.",
         "unsupported",
         "Evidence describes adversarial training, not supervised cross-entropy."),
        ("I only found Main Idea mentioned in profile/course context.",
         "(any evidence snippet)",
         "unsupported",
         "System-level abstention — never supported."),
        ("AlphaFold predicts protein structures from amino-acid sequences.",
         "We developed AlphaFold, a neural network that predicts 3D protein structures with atomic accuracy.",
         "supported",
         "Evidence directly asserts the claim."),
        ("LLaMA 2 reaches GPT-4 performance across most benchmarks.",
         "LLaMA 2 is competitive with closed-source chat models on a range of tasks.",
         "unsupported",
         "Evidence says 'competitive,' not 'reaches GPT-4' — claim overstates."),
    ]
    for idx, (claim, ev, label, why) in enumerate(examples, start=9):
        ws2.cell(row=idx, column=1, value=claim)
        ws2.cell(row=idx, column=2, value=ev)
        ws2.cell(row=idx, column=3, value=label)
        ws2.cell(row=idx, column=4, value=why)
    for col_letter, width in [("A", 55), ("B", 80), ("C", 15), ("D", 50)]:
        ws2.column_dimensions[col_letter].width = width
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 3. Paper List
    ws3 = wb.create_sheet("Paper List")
    ws3["A1"] = "Round-2 Corpus (15 diverse papers)"
    ws3["A1"].font = Font(bold=True, size=12)
    headers = ["#", "File", "Paper title", "Field", "Venue / Year", "Official link"]
    for i, h in enumerate(headers, start=1):
        ws3.cell(row=3, column=i, value=h).font = Font(bold=True)
    for i, (fname, title, field, venue, link) in enumerate(PAPER_LIST, start=4):
        ws3.cell(row=i, column=1, value=i - 3)
        ws3.cell(row=i, column=2, value=fname)
        ws3.cell(row=i, column=3, value=title)
        ws3.cell(row=i, column=4, value=field)
        ws3.cell(row=i, column=5, value=venue)
        link_cell = ws3.cell(row=i, column=6, value=link)
        link_cell.hyperlink = link
        link_cell.font = Font(color="0563C1", underline="single")
    for col, width in [("A", 5), ("B", 26), ("C", 75), ("D", 24), ("E", 30), ("F", 50)]:
        ws3.column_dimensions[col].width = width

    # 4. Labeling
    ws4 = wb.create_sheet("Labeling")
    ws4.freeze_panes = "A2"
    headers = [
        "#", "query_id", "query", "mode", "query_type",
        "claim_id", "claim_text", "evidence_text",
        f"Your Label (Coder {coder})", "Notes",
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws4.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")
    for idx, p in enumerate(pairs, start=1):
        ws4.cell(row=idx + 1, column=1, value=idx)
        ws4.cell(row=idx + 1, column=2, value=_xlsx_safe(p.get("query_id")))
        ws4.cell(row=idx + 1, column=3, value=_xlsx_safe(p.get("query")))
        ws4.cell(row=idx + 1, column=4, value=_xlsx_safe(p.get("mode")))
        ws4.cell(row=idx + 1, column=5, value=_xlsx_safe(p.get("query_type")))
        ws4.cell(row=idx + 1, column=6, value=_xlsx_safe(p.get("claim_id")))
        ws4.cell(row=idx + 1, column=7, value=_xlsx_safe(p.get("claim_text")))
        ws4.cell(row=idx + 1, column=8, value=_xlsx_safe(p.get("evidence_text")))
        ws4.cell(row=idx + 1, column=9, value="")  # label — to fill via dropdown
        ws4.cell(row=idx + 1, column=10, value="")  # notes
    # Dropdown validation on column I (label)
    dv = DataValidation(
        type="list",
        formula1='"supported,unsupported"',
        allow_blank=False,
        showDropDown=False,
        error="Use the dropdown only",
        errorTitle="Invalid label",
    )
    dv.add(f"I2:I{len(pairs) + 1}")
    ws4.add_data_validation(dv)
    widths = {"A": 5, "B": 8, "C": 55, "D": 10, "E": 14, "F": 16, "G": 60, "H": 75, "I": 18, "J": 35}
    for c, w in widths.items():
        ws4.column_dimensions[c].width = w
    for row in ws4.iter_rows(min_row=2, max_row=len(pairs) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 5. Summary
    ws5 = wb.create_sheet("Summary")
    ws5["A1"] = f"Progress — Coder {coder}"
    ws5["A1"].font = Font(bold=True, size=12)
    total = len(pairs)
    label_col = f"Labeling!I2:I{total + 1}"
    ws5["A3"] = "Total rows"
    ws5["B3"] = total
    ws5["A4"] = "Labeled"
    ws5["B4"] = f"=COUNTIF({label_col},\"<>\")"
    ws5["A5"] = "Remaining"
    ws5["B5"] = f"={total}-COUNTIF({label_col},\"<>\")"
    ws5["A7"] = "supported"
    ws5["B7"] = f"=COUNTIF({label_col},\"supported\")"
    ws5["A8"] = "unsupported"
    ws5["B8"] = f"=COUNTIF({label_col},\"unsupported\")"
    ws5["A10"] = "By mode"
    ws5["A10"].font = Font(bold=True)
    ws5["A11"] = "uploaded (labeled)"
    ws5["B11"] = f"=COUNTIFS(Labeling!D2:D{total + 1},\"uploaded\",{label_col},\"<>\")"
    ws5["A12"] = "public (labeled)"
    ws5["B12"] = f"=COUNTIFS(Labeling!D2:D{total + 1},\"public\",{label_col},\"<>\")"
    for c, w in [("A", 26), ("B", 18)]:
        ws5.column_dimensions[c].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    t0 = time.time()
    pairs_path = Path("Evaluation/queries/claim_evidence_pairs.json")
    include_public = os.getenv("CODEBOOK_INCLUDE_PUBLIC", "false").strip().lower() in {"1", "true", "yes"}
    # Default 120s per-call timeout — public-mode queries hit external APIs
    # (arXiv, OpenAlex, Semantic Scholar, CrossRef) which can take 30-60s per
    # call when CrossRef is rate-limited.
    per_call_timeout = int(os.getenv("CODEBOOK_PER_CALL_TIMEOUT", "120"))
    all_pairs = run_pipeline(pairs_path, include_public=include_public, per_call_timeout_s=per_call_timeout)
    print()
    print(f"Collected {len(all_pairs)} pairs in {int(time.time() - t0)}s")
    print(f"Pairs dumped to -> {pairs_path}")

    for coder in CODERS:
        out = OUT_DIR / f"ScholarRAG_Unified_Labeling_Coder_{coder}.xlsx"
        write_codebook(out, coder, all_pairs)
        print(f"  wrote {out}")

    print()
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
