"""Fit the unified MSA logistic on human-labeled gold data and run the per-mode ablation.

Reads:  Evaluation/data/calibration/features.xlsx   (claim_id, gold_label, M, S, A, mode)
Writes: Evaluation/data/calibration/calibration_fit.json   (weights, metrics, ablation)
        Evaluation/data/calibration/reliability_diagram.xlsx  (bucket-mean P / empirical rate)

Also writes the winning weights to the `confidence_calibration` DB table under
`label='unified'` when --write-db is passed.

Usage:
    python -m backend.scripts.fit_unified_calibration
    python -m backend.scripts.fit_unified_calibration --write-db
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List

import numpy as np
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import brier_score_loss, log_loss, roc_auc_score

from backend.services.db import fetchone

FEATURES_PATH = Path("Evaluation/data/calibration/features.xlsx")
OUT_JSON = Path("Evaluation/data/calibration/calibration_fit.json")
OUT_DIAGRAM = Path("Evaluation/data/calibration/reliability_diagram.xlsx")


def load_features() -> List[Dict]:
    if not FEATURES_PATH.exists():
        print(f"[!] Missing {FEATURES_PATH}. Run extract_msa_features.py first.")
        sys.exit(1)
    from openpyxl import load_workbook
    wb = load_workbook(FEATURES_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    headers = [str(h or "") for h in next(it, [])]
    idx = {h: i for i, h in enumerate(headers)}
    rows: List[Dict] = []
    for raw in it:
        rows.append({
            "claim_id": raw[idx["claim_id"]],
            "query_id": raw[idx["query_id"]],
            "mode": raw[idx["mode"]],
            "query_type": raw[idx["query_type"]],
            "gold_label": raw[idx["gold_label"]],
            "M": float(raw[idx["M"]] or 0.0),
            "S": float(raw[idx["S"]] or 0.0),
            "A": float(raw[idx["A"]] or 0.0),
        })
    return rows


def fit_logistic(rows: List[Dict], label: str = "all") -> Dict:
    """Fit sigmoid(b + w1*M + w2*S + w3*A) with sklearn.

    Returns dict with weights, Brier, log-loss, AUC, n.
    """
    X = np.array([[r["M"], r["S"], r["A"]] for r in rows], dtype=float)
    y = np.array([1 if r["gold_label"] == "supported" else 0 for r in rows], dtype=int)

    if len(set(y)) < 2:
        return {"label": label, "n": len(rows), "error": "single-class labels, cannot fit"}

    # No regularization — we want plain MLE; penalty='l2' is sklearn default but C=1e9 disables it.
    lr = LogisticRegression(C=1e9, solver="lbfgs", max_iter=500)
    lr.fit(X, y)

    coefs = lr.coef_[0]
    bias = float(lr.intercept_[0])
    w = {
        "w1": round(float(coefs[0]), 4),
        "w2": round(float(coefs[1]), 4),
        "w3": round(float(coefs[2]), 4),
        "b": round(bias, 4),
    }

    preds = lr.predict_proba(X)[:, 1]
    try:
        brier = brier_score_loss(y, preds)
    except Exception:
        brier = None
    try:
        ll = log_loss(y, preds, labels=[0, 1])
    except Exception:
        ll = None
    try:
        auc = roc_auc_score(y, preds) if len(set(y)) == 2 else None
    except Exception:
        auc = None

    positives = int(y.sum())
    return {
        "label": label,
        "n": len(rows),
        "n_supported": positives,
        "n_unsupported": len(rows) - positives,
        "base_rate_supported": round(positives / max(1, len(rows)), 4),
        "weights": w,
        "brier": round(float(brier), 6) if brier is not None else None,
        "log_loss": round(float(ll), 6) if ll is not None else None,
        "auc": round(float(auc), 4) if auc is not None else None,
    }


def reliability_diagram(rows: List[Dict], weights: Dict, n_bins: int = 10) -> List[Dict]:
    """Bucket predictions into deciles, report mean-pred vs empirical supported-rate."""
    X = np.array([[r["M"], r["S"], r["A"]] for r in rows], dtype=float)
    y = np.array([1 if r["gold_label"] == "supported" else 0 for r in rows], dtype=int)
    w1, w2, w3, b = weights["w1"], weights["w2"], weights["w3"], weights["b"]
    logits = b + w1 * X[:, 0] + w2 * X[:, 1] + w3 * X[:, 2]
    probs = 1.0 / (1.0 + np.exp(-logits))

    bins = np.linspace(0.0, 1.0, n_bins + 1)
    buckets: List[Dict] = []
    for i in range(n_bins):
        lo, hi = bins[i], bins[i + 1]
        if i == n_bins - 1:
            mask = (probs >= lo) & (probs <= hi)
        else:
            mask = (probs >= lo) & (probs < hi)
        n = int(mask.sum())
        if n == 0:
            continue
        buckets.append({
            "bucket_lo": round(float(lo), 4),
            "bucket_hi": round(float(hi), 4),
            "n": n,
            "mean_predicted": round(float(probs[mask].mean()), 4),
            "empirical_supported_rate": round(float(y[mask].mean()), 4),
        })
    return buckets


def write_to_db(weights: Dict, label: str, metrics: Dict) -> None:
    row = fetchone(
        """
        INSERT INTO confidence_calibration (label, weights, metrics, dataset_size, model_name, created_at)
        VALUES (%s, %s, %s, %s, %s, now())
        RETURNING id
        """,
        [label, json.dumps(weights), json.dumps(metrics), metrics.get("n", 0), "logistic_v2"],
    )
    print(f"Inserted DB row id={row['id']} label={label!r}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--write-db", action="store_true", help="Write winning weights to confidence_calibration table")
    args = parser.parse_args()

    rows = load_features()
    print(f"Loaded {len(rows)} rows from {FEATURES_PATH}")
    from collections import Counter
    dist = Counter(r["gold_label"] for r in rows)
    print(f"Label distribution: {dict(dist)}")
    modes = Counter(r["mode"] for r in rows)
    print(f"Mode distribution:  {dict(modes)}")

    # --- Three-way ablation ---
    print("\n=== Ablation: pooled vs uploaded-only vs public-only ===")
    pooled = fit_logistic(rows, label="pooled")
    uploaded = fit_logistic([r for r in rows if r["mode"] == "uploaded"], label="uploaded_only")
    public = fit_logistic([r for r in rows if r["mode"] == "public"], label="public_only")

    for name, fit in [("pooled", pooled), ("uploaded", uploaded), ("public", public)]:
        print(f"\n-- {name} --")
        if "error" in fit:
            print(f"  ERROR: {fit['error']}")
            continue
        print(f"  n={fit['n']}  supported={fit['n_supported']}  base_rate={fit['base_rate_supported']:.3f}")
        w = fit["weights"]
        print(f"  weights: w1(M)={w['w1']:.3f}  w2(S)={w['w2']:.3f}  w3(A)={w['w3']:.3f}  b={w['b']:.3f}")
        print(f"  Brier={fit['brier']}  log-loss={fit['log_loss']}  AUC={fit['auc']}")

    # --- Decide which to use ---
    unified_valid = False
    comparison = {}
    if "error" not in uploaded and "error" not in public and "error" not in pooled:
        # Compare pooled against per-mode: if pooled Brier is within ~0.02 of the per-mode average,
        # treat pooled as statistically valid.
        per_mode_avg = ((uploaded["brier"] or 0) * uploaded["n"] + (public["brier"] or 0) * public["n"]) / max(1, (uploaded["n"] + public["n"]))
        comparison = {
            "pooled_brier": pooled["brier"],
            "per_mode_avg_brier": round(per_mode_avg, 6),
            "delta_pooled_minus_permode": round((pooled["brier"] or 0) - per_mode_avg, 6),
            "pooled_is_unified_valid": bool((pooled["brier"] or 0) - per_mode_avg < 0.02),
        }
        unified_valid = comparison["pooled_is_unified_valid"]
        print()
        print(f"Pooled Brier:       {pooled['brier']}")
        print(f"Per-mode avg Brier: {round(per_mode_avg, 4)}")
        print(f"Delta:              {round((pooled['brier'] or 0) - per_mode_avg, 4)}")
        print(f"Unified valid:      {unified_valid}")

    # --- Reliability diagram from pooled fit ---
    diagram = reliability_diagram(rows, pooled["weights"]) if "error" not in pooled else []
    if diagram:
        print("\n=== Reliability diagram (pooled fit, 10 buckets) ===")
        print(f"  {'bucket':<20}{'n':>5}  mean_pred  empirical")
        for b in diagram:
            print(f"  [{b['bucket_lo']:.2f}, {b['bucket_hi']:.2f}){b['n']:>6}    "
                  f"{b['mean_predicted']:.3f}     {b['empirical_supported_rate']:.3f}")

    # --- Save ---
    out = {
        "pooled": pooled,
        "uploaded_only": uploaded,
        "public_only": public,
        "ablation": comparison,
        "unified_valid": unified_valid,
    }
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nWrote fit report → {OUT_JSON}")

    if diagram:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        fields = ["bucket_lo", "bucket_hi", "n", "mean_predicted", "empirical_supported_rate"]
        wb = Workbook()
        ws = wb.active
        ws.title = "reliability"
        ws.append(fields)
        for r in diagram:
            ws.append([r.get(f, "") for f in fields])
        ws.freeze_panes = "A2"
        widths = {"bucket_lo": 11, "bucket_hi": 11, "n": 8,
                  "mean_predicted": 16, "empirical_supported_rate": 24}
        for i, f in enumerate(fields, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(f, 12)
        wb.save(OUT_DIAGRAM)
        print(f"Wrote reliability diagram → {OUT_DIAGRAM}")

    # --- DB write ---
    if args.write_db and "error" not in pooled:
        print("\n=== Writing to DB ===")
        write_to_db(pooled["weights"], "unified", {
            "n": pooled["n"],
            "brier": pooled["brier"],
            "log_loss": pooled["log_loss"],
            "auc": pooled["auc"],
            "unified_valid": unified_valid,
            "ablation": comparison,
        })

    return 0


if __name__ == "__main__":
    sys.exit(main())
