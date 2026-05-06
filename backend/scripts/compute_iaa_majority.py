"""Compute inter-annotator agreement + majority-vote resolve human labels.

Reads the three filled codebooks from Downloads/, writes:
    - Evaluation/data/calibration/iaa_report.json         (pairwise kappa + overall)
    - Evaluation/data/calibration/gold_labels.xlsx        (majority-vote resolved)
    - Evaluation/data/calibration/label_distribution.json (aggregate stats)

Usage:
    python -m backend.scripts.compute_iaa_majority
"""

from __future__ import annotations

import json
import os
import random
import sys
from collections import Counter, defaultdict
from itertools import combinations
from pathlib import Path
from typing import Dict, Iterable, List, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Target dataset size: stratified-sample this many pairs from the full coder
# workbook output (the labelers may have produced more). Set to 0 or unset to
# keep all pairs.
TARGET_N = int(os.getenv("CALIBRATION_TARGET_N", "530"))
RNG_SEED = int(os.getenv("CALIBRATION_SAMPLE_SEED", "42"))


def _write_xlsx(path: Path, fields: List[str], rows: Iterable[Mapping], sheet_name: str = "data") -> None:
    """Write a list of dicts to xlsx with frozen header row + readable column widths."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(fields)
    body: list[list] = []
    for r in rows:
        vals = [r.get(f, "") for f in fields]
        ws.append(vals)
        body.append(vals)
    ws.freeze_panes = "A2"
    for i, col_values in enumerate(zip(fields, *body), start=1):
        max_len = max((len(str(v)) for v in col_values), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, max_len + 2))
    wb.save(path)


_WORKBOOKS_DIR = Path("Evaluation/data/calibration/coder_workbooks")
INPUTS = {
    "A": str(_WORKBOOKS_DIR / "coder_A.xlsx"),
    "B": str(_WORKBOOKS_DIR / "coder_B.xlsx"),
    "C": str(_WORKBOOKS_DIR / "coder_C.xlsx"),
}

OUT_DIR = Path("Evaluation/data/calibration")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def _clean(v) -> str:
    s = str(v or "").strip().lower()
    if s in {"supported", "unsupported"}:
        return s
    return ""


def load_codebook(path: str) -> List[Dict]:
    wb = load_workbook(path)
    ws = wb["Labeling"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append(
            {
                "row": r,
                "claim_id": ws.cell(row=r, column=idx["claim_id"]).value,
                "query_id": ws.cell(row=r, column=idx["query_id"]).value,
                "mode": (ws.cell(row=r, column=idx["mode"]).value or "").strip(),
                "query_type": (ws.cell(row=r, column=idx["query_type"]).value or "").strip(),
                "claim_text": ws.cell(row=r, column=idx["claim_text"]).value or "",
                "evidence_text": ws.cell(row=r, column=idx["evidence_text"]).value or "",
                "label": _clean(ws.cell(row=r, column=9).value),
                "notes": (ws.cell(row=r, column=10).value or "").strip(),
            }
        )
    return rows


def cohens_kappa(labels_1: List[str], labels_2: List[str]) -> Dict:
    """Cohen's kappa for two raters with the same categorical labels."""
    assert len(labels_1) == len(labels_2)
    n = len(labels_1)
    cats = sorted({*labels_1, *labels_2})
    # Confusion
    confusion = {a: Counter() for a in cats}
    for a, b in zip(labels_1, labels_2):
        confusion[a][b] += 1
    # Observed agreement
    po = sum(confusion[a][a] for a in cats) / n
    # Expected agreement
    p1 = {a: sum(confusion[a].values()) / n for a in cats}
    p2 = {b: sum(confusion[a][b] for a in cats) / n for b in cats}
    pe = sum(p1[a] * p2[a] for a in cats)
    kappa = (po - pe) / (1 - pe) if pe < 1 else 1.0
    return {
        "po": round(po, 4),
        "pe": round(pe, 4),
        "kappa": round(kappa, 4),
        "n": n,
        "confusion": {a: dict(confusion[a]) for a in cats},
    }


def _interpret_kappa(k: float) -> str:
    if k < 0.0:
        return "poor (worse than chance)"
    if k < 0.2:
        return "slight"
    if k < 0.4:
        return "fair"
    if k < 0.6:
        return "moderate"
    if k < 0.8:
        return "substantial"
    return "almost perfect"


def _stratified_sample(indices: list[int], strata_keys: list[tuple], target_n: int, seed: int) -> list[int]:
    """Stratified sample target_n indices, preserving each stratum's proportion.

    `strata_keys[i]` is the stratum key for `indices[i]`. Returns the indices
    to keep, sorted. Uses `random.Random(seed)` for reproducibility.
    """
    rng = random.Random(seed)
    bucket: dict[tuple, list[int]] = defaultdict(list)
    for idx, key in zip(indices, strata_keys):
        bucket[key].append(idx)
    total = len(indices)
    kept: list[int] = []
    for key, ids in bucket.items():
        share = len(ids) / total
        take = int(round(share * target_n))
        take = min(take, len(ids))
        rng.shuffle(ids)
        kept.extend(ids[:take])
    # Correct rounding error (±1) to hit target_n exactly.
    while len(kept) > target_n:
        kept.pop()
    while len(kept) < target_n and len(kept) < total:
        remaining = [i for i in indices if i not in set(kept)]
        rng.shuffle(remaining)
        kept.extend(remaining[: target_n - len(kept)])
    return sorted(kept)


def main() -> int:
    # Load all three
    data = {coder: load_codebook(path) for coder, path in INPUTS.items()}

    # Align by row index
    n_full = len(data["A"])
    assert all(len(data[c]) == n_full for c in data), "Row count mismatch between coders"

    # Confirm claim_ids align
    for i in range(n_full):
        ids = {c: data[c][i]["claim_id"] for c in data}
        if len(set(ids.values())) > 1:
            raise RuntimeError(f"claim_id mismatch at row {i + 2}: {ids}")

    print(f"Loaded {n_full} rows per coder, aligned by claim_id.")

    # Stratified downsample to TARGET_N (default 530) to match the stated
    # calibration-dataset size. Strata = (mode, query_type). Deterministic
    # via RNG_SEED so the same rows are selected every run.
    if TARGET_N > 0 and TARGET_N < n_full:
        strata = [(data["A"][i]["mode"], data["A"][i]["query_type"]) for i in range(n_full)]
        keep = _stratified_sample(list(range(n_full)), strata, TARGET_N, RNG_SEED)
        keep_set = set(keep)
        for coder in data:
            data[coder] = [row for i, row in enumerate(data[coder]) if i in keep_set]
        print(f"Stratified-sampled {len(keep)} of {n_full} pairs (seed={RNG_SEED}, strata=mode×query_type).")
    n = len(data["A"])

    # Pairwise kappa
    iaa = {"n_pairs": n, "pairwise": {}}
    for a, b in combinations(sorted(data.keys()), 2):
        la = [row["label"] for row in data[a]]
        lb = [row["label"] for row in data[b]]
        k = cohens_kappa(la, lb)
        k["interpretation"] = _interpret_kappa(k["kappa"])
        iaa["pairwise"][f"{a}-{b}"] = k
        print(f"Cohen's kappa {a}-{b}: {k['kappa']:.4f} ({k['interpretation']}, po={k['po']:.3f}, pe={k['pe']:.3f})")

    # Average of the three pairwise kappas
    avg_kappa = sum(iaa["pairwise"][p]["kappa"] for p in iaa["pairwise"]) / len(iaa["pairwise"])
    iaa["average_pairwise_kappa"] = round(avg_kappa, 4)
    iaa["average_interpretation"] = _interpret_kappa(avg_kappa)
    print(f"Average pairwise kappa: {avg_kappa:.4f} ({_interpret_kappa(avg_kappa)})")

    # Per-coder distribution
    iaa["per_coder_distribution"] = {c: dict(Counter(row["label"] for row in data[c])) for c in data}

    # Majority-vote resolution
    gold_rows = []
    disagreement_count = 0
    three_way_tie = 0
    for i in range(n):
        labels = [data[c][i]["label"] for c in ("A", "B", "C")]
        cnt = Counter(labels)
        # Majority = most common label among labelled ones
        # With 3 binary labels, you always have a majority winner.
        top, top_count = cnt.most_common(1)[0]
        if top_count == 3:
            resolution = "unanimous"
        elif top_count == 2:
            resolution = "majority"
            disagreement_count += 1
        else:
            # Three different? Impossible with binary. But handle just in case.
            resolution = "tie"
            three_way_tie += 1
        base = data["A"][i]  # any coder's row; metadata is identical
        gold_rows.append(
            {
                "claim_id": base["claim_id"],
                "query_id": base["query_id"],
                "mode": base["mode"],
                "query_type": base["query_type"],
                "claim_text": base["claim_text"],
                "evidence_text": base["evidence_text"],
                "label_A": data["A"][i]["label"],
                "label_B": data["B"][i]["label"],
                "label_C": data["C"][i]["label"],
                "gold_label": top,
                "resolution": resolution,
            }
        )

    iaa["majority_vote_summary"] = {
        "unanimous": sum(1 for r in gold_rows if r["resolution"] == "unanimous"),
        "majority_2_of_3": sum(1 for r in gold_rows if r["resolution"] == "majority"),
        "three_way_tie": three_way_tie,
    }

    gold_dist = Counter(r["gold_label"] for r in gold_rows)
    iaa["gold_distribution"] = dict(gold_dist)

    # By mode + query type
    iaa["gold_by_mode"] = {}
    for m in ("uploaded", "public"):
        subset = [r for r in gold_rows if r["mode"] == m]
        iaa["gold_by_mode"][m] = {
            "total": len(subset),
            "supported": sum(1 for r in subset if r["gold_label"] == "supported"),
            "unsupported": sum(1 for r in subset if r["gold_label"] == "unsupported"),
        }

    iaa["gold_by_query_type"] = {}
    for t in ("definitional", "methodology", "factual", "limitations"):
        subset = [r for r in gold_rows if r["query_type"] == t]
        iaa["gold_by_query_type"][t] = {
            "total": len(subset),
            "supported": sum(1 for r in subset if r["gold_label"] == "supported"),
            "unsupported": sum(1 for r in subset if r["gold_label"] == "unsupported"),
        }

    # Emit
    iaa_path = OUT_DIR / "iaa_report.json"
    iaa_path.write_text(json.dumps(iaa, indent=2))
    print(f"\nWrote IAA report → {iaa_path}")

    gold_path = OUT_DIR / "gold_labels.xlsx"
    fields = [
        "claim_id",
        "query_id",
        "mode",
        "query_type",
        "claim_text",
        "evidence_text",
        "label_A",
        "label_B",
        "label_C",
        "gold_label",
        "resolution",
    ]
    _write_xlsx(gold_path, fields, gold_rows, sheet_name="gold_labels")
    print(f"Wrote gold labels → {gold_path}")

    dist_path = OUT_DIR / "label_distribution.json"
    dist_path.write_text(
        json.dumps(
            {
                "n": n,
                "per_coder": iaa["per_coder_distribution"],
                "gold": iaa["gold_distribution"],
                "gold_by_mode": iaa["gold_by_mode"],
                "gold_by_query_type": iaa["gold_by_query_type"],
            },
            indent=2,
        )
    )
    print(f"Wrote distribution → {dist_path}")

    print()
    print("=== HEADLINE STATS ===")
    print(f"Total claims:              {n}")
    print(f"Average pairwise kappa:    {avg_kappa:.4f} ({_interpret_kappa(avg_kappa)})")
    print(
        f"Unanimous agreement:       {iaa['majority_vote_summary']['unanimous']} / {n}  ({iaa['majority_vote_summary']['unanimous'] / n:.1%})"
    )
    print(f"Gold label distribution:   {dict(gold_dist)}")
    print(f"Gold supported rate:       {gold_dist['supported'] / n:.1%}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
